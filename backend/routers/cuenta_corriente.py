from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import text
from core.database import get_db
from models.models import Cliente, MovimientoCuenta, Venta

router = APIRouter(prefix="/api/v1/cuenta-corriente", tags=["cuenta-corriente"])


@router.get("/clientes")
def listar_clientes_cta_cte(solo_con_deuda: bool = True, db: Session = Depends(get_db)):
    q = db.query(Cliente)
    if solo_con_deuda:
        q = q.filter(Cliente.saldo_deudor > 0)
    clientes = q.order_by(Cliente.nombre).all()
    return {
        "data": [
            {
                "id": c.id,
                "nombre": c.nombre,
                "telefono": c.telefono or "",
                "dni_cuit": c.dni_cuit or "",
                "tipo_cliente": c.tipo_cliente or "persona",
                "saldo_deudor": c.saldo_deudor or 0,
            }
            for c in clientes
        ]
    }


@router.get("/{cliente_id}/movimientos")
def movimientos_cliente(cliente_id: int, db: Session = Depends(get_db)):
    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    movs = (
        db.query(MovimientoCuenta)
        .filter(MovimientoCuenta.cliente_id == cliente_id)
        .order_by(MovimientoCuenta.fecha.desc())
        .limit(100)
        .all()
    )

    return {
        "cliente": {
            "id": cliente.id,
            "nombre": cliente.nombre,
            "telefono": cliente.telefono or "",
            "saldo_deudor": cliente.saldo_deudor or 0,
        },
        "movimientos": [
            {
                "id": m.id,
                "tipo": m.tipo,
                "monto": m.monto,
                "descripcion": m.descripcion,
                "metodo_pago": m.metodo_pago or "",
                "fecha": m.fecha.isoformat() if m.fecha else None,
            }
            for m in movs
        ],
    }


@router.post("/migrar-descripciones")
def migrar_descripciones(db: Session = Depends(get_db)):
    """Corrige los movimientos que dicen 'Venta #' → 'Factura N°'."""
    from sqlalchemy import text
    db.execute(text(
        "UPDATE movimientos_cuenta SET descripcion = REPLACE(descripcion, 'Venta #', 'Factura N° ') "
        "WHERE descripcion LIKE 'Venta #%'"
    ))
    db.commit()
    return {"ok": True}


@router.get("/{cliente_id}/facturas-adeudadas")
def facturas_adeudadas(cliente_id: int, db: Session = Depends(get_db)):
    """Devuelve las ventas con monto_debe > 0 del cliente."""
    ventas = (
        db.query(Venta)
        .filter(
            Venta.cliente_id == cliente_id,
            Venta.monto_debe > 0,
            Venta.es_cotizacion == False,
        )
        .order_by(Venta.fecha_creacion.asc())
        .all()
    )
    return [
        {
            "id": v.id,
            "fecha": v.fecha_creacion.strftime("%d/%m/%Y") if v.fecha_creacion else "-",
            "total": v.total_venta or 0,
            "abonado": v.monto_abonado or 0,
            "debe": v.monto_debe or 0,
        }
        for v in ventas
    ]


@router.post("/pagar-imputado")
def pagar_imputado(data: dict, db: Session = Depends(get_db)):
    """Imputa un pago contra facturas específicas del cliente."""
    cliente_id = data.get("cliente_id")
    imputaciones = data.get("imputaciones", [])  # [{venta_id, monto}]
    metodo = data.get("metodo_pago", "efectivo")
    observaciones = data.get("observaciones", "")

    if not cliente_id or not imputaciones:
        raise HTTPException(status_code=400, detail="Datos inválidos")

    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    total_pagado = 0
    for imp in imputaciones:
        venta_id = imp.get("venta_id")
        monto = float(imp.get("monto", 0))
        if monto <= 0:
            continue
        venta = db.query(Venta).filter(Venta.id == venta_id, Venta.cliente_id == cliente_id).first()
        if not venta:
            continue
        monto_aplicar = min(monto, venta.monto_debe)
        venta.monto_abonado = (venta.monto_abonado or 0) + monto_aplicar
        venta.monto_debe = max(0, (venta.monto_debe or 0) - monto_aplicar)
        total_pagado += monto_aplicar

        mov = MovimientoCuenta(
            cliente_id=cliente_id,
            tipo="pago",
            monto=monto_aplicar,
            descripcion=f"Pago Factura N° {venta_id}" + (f" - {observaciones}" if observaciones else ""),
            metodo_pago=metodo,
        )
        db.add(mov)

    cliente.saldo_deudor = max(0, (cliente.saldo_deudor or 0) - total_pagado)
    db.commit()

    return {"message": "Pago imputado", "total_pagado": total_pagado, "nuevo_saldo": cliente.saldo_deudor}


@router.post("/pagar")
def registrar_pago(data: dict, db: Session = Depends(get_db)):
    cliente_id = data.get("cliente_id")
    monto = data.get("monto", 0)
    metodo = data.get("metodo_pago", "efectivo")
    observaciones = data.get("observaciones", "")

    if not cliente_id or monto <= 0:
        raise HTTPException(status_code=400, detail="Datos inválidos")

    cliente = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    # Registrar movimiento
    mov = MovimientoCuenta(
        cliente_id=cliente_id,
        tipo="pago",
        monto=monto,
        descripcion=observaciones or f"Pago {metodo}",
        metodo_pago=metodo,
    )
    db.add(mov)

    # Actualizar saldo
    cliente.saldo_deudor = max(0, (cliente.saldo_deudor or 0) - monto)
    db.commit()

    return {
        "message": "Pago registrado",
        "nuevo_saldo": cliente.saldo_deudor,
    }


@router.post("/migrar-columnas-clientes")
def migrar_columnas_clientes(db: Session = Depends(get_db)):
    """Agrega tipo_cliente y correo a la tabla clientes si no existen."""
    try:
        db.execute(text("ALTER TABLE clientes ADD COLUMN tipo_cliente VARCHAR DEFAULT 'persona'"))
        db.commit()
    except Exception:
        db.rollback()
    try:
        db.execute(text("ALTER TABLE clientes ADD COLUMN correo VARCHAR DEFAULT ''"))
        db.commit()
    except Exception:
        db.rollback()
    return {"ok": True}


@router.post("/importar-contactos")
def importar_contactos(data: dict, db: Session = Depends(get_db)):
    """
    Importa contactos desde el Excel.
    data: { contactos: [{tipo, dni_cuit, nombre, telefono, correo}] }
    Hace upsert por nombre (case-insensitive). Devuelve contadores.
    """
    contactos = data.get("contactos", [])
    if not contactos:
        raise HTTPException(status_code=400, detail="Sin contactos")

    creados = 0
    actualizados = 0

    for c in contactos:
        nombre = (c.get("nombre") or "").strip()
        if not nombre:
            continue

        dni = (c.get("dni_cuit") or "").strip()
        telefono = (c.get("telefono") or "").strip()
        correo = (c.get("correo") or "").strip()
        tipo = (c.get("tipo") or "persona").strip().lower()
        if tipo not in ("persona", "empresa"):
            tipo = "persona"

        # Buscar por DNI/CUIT primero, luego por nombre
        existente = None
        if dni and dni.upper() not in ("", "NOSE", "..", "0"):
            existente = db.query(Cliente).filter(Cliente.dni_cuit == dni).first()
        if not existente:
            existente = db.query(Cliente).filter(
                text("LOWER(nombre) = LOWER(:n)")
            ).params(n=nombre).first()

        if existente:
            existente.tipo_cliente = tipo
            if dni:
                existente.dni_cuit = dni
            if telefono:
                existente.telefono = telefono
            if correo:
                existente.correo = correo
            actualizados += 1
        else:
            nuevo = Cliente(
                nombre=nombre,
                dni_cuit=dni,
                telefono=telefono,
                correo=correo,
                tipo_cliente=tipo,
                saldo_deudor=0,
                activo=True,
            )
            db.add(nuevo)
            creados += 1

    db.commit()
    return {"ok": True, "creados": creados, "actualizados": actualizados}


@router.post("/importar-contactos-lote")
def importar_contactos_lote(data: dict, db: Session = Depends(get_db)):
    """Procesa un lote (offset+limit) del Excel en el servidor. Para evitar timeout."""
    import os, openpyxl, re

    offset = int(data.get("offset", 0))
    limit = int(data.get("limit", 50))

    base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    posibles = [
        os.path.join(base, "frontend", "admin", "catalogos", "contactos.xlsx"),
        os.path.join(base, "contactos.xlsx"),
    ]
    ruta = next((p for p in posibles if os.path.exists(p)), None)
    if not ruta:
        raise HTTPException(status_code=404, detail="contactos.xlsx no encontrado")

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.active
    # Saltar header (fila 0) + offset
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    total = len(all_rows) - 1  # sin header
    lote = all_rows[1 + offset: 1 + offset + limit]

    creados = 0
    actualizados = 0

    for row in lote:
        tipo_raw = (str(row[0] or "")).strip().upper()
        tipo = "empresa" if tipo_raw == "EMPRESA" else "persona"
        dni_raw = (str(row[1] or "")).strip()
        apellido = (str(row[2] or "")).strip()
        nombre_raw = (str(row[3] or "")).strip()
        wa = (str(row[4] or "")).strip()
        correo_raw = (str(row[5] or "")).strip() if len(row) > 5 else ""

        nombre = f"{apellido} {nombre_raw}".strip() if tipo == "persona" and apellido else nombre_raw
        if not nombre or nombre in ("None None", "None", ""):
            continue

        dni = dni_raw if re.match(r"^\d{6,}$", dni_raw) else ""
        telefono = wa if re.search(r"\d{6,}", wa) else ""
        correo = correo_raw if correo_raw != "None" else ""

        existente = None
        if dni:
            existente = db.query(Cliente).filter(Cliente.dni_cuit == dni).first()
        if not existente:
            existente = db.query(Cliente).filter(
                text("LOWER(nombre) = LOWER(:n)")
            ).params(n=nombre).first()

        if existente:
            existente.tipo_cliente = tipo
            if dni: existente.dni_cuit = dni
            if telefono: existente.telefono = telefono
            if correo: existente.correo = correo
            actualizados += 1
        else:
            db.add(Cliente(nombre=nombre, dni_cuit=dni, telefono=telefono,
                           correo=correo, tipo_cliente=tipo, saldo_deudor=0, activo=True))
            creados += 1

    db.commit()
    return {"ok": True, "creados": creados, "actualizados": actualizados,
            "total": total, "procesados": offset + len(lote), "fin": offset + len(lote) >= total}


@router.post("/importar-contactos-servidor")
def importar_contactos_servidor(db: Session = Depends(get_db)):
    """Lee contactos.xlsx del filesystem del servidor y hace upsert en clientes."""
    import os
    import openpyxl

    # Buscar el archivo en ubicaciones posibles
    base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    posibles = [
        os.path.join(base, "frontend", "admin", "catalogos", "contactos.xlsx"),
        os.path.join(base, "contactos.xlsx"),
        "/tmp/contactos.xlsx",
    ]
    ruta = next((p for p in posibles if os.path.exists(p)), None)
    if not ruta:
        raise HTTPException(status_code=404, detail="contactos.xlsx no encontrado en el servidor")

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    creados = 0
    actualizados = 0

    for row in rows[1:]:  # saltar header
        tipo_raw = (str(row[0] or "")).strip().upper()
        tipo = "empresa" if tipo_raw == "EMPRESA" else "persona"
        dni_raw = (str(row[1] or "")).strip()
        apellido = (str(row[2] or "")).strip()
        nombre_raw = (str(row[3] or "")).strip()
        wa = (str(row[4] or "")).strip()
        correo = (str(row[5] or "")).strip() if len(row) > 5 else ""

        # Ignorar filas sin nombre
        if tipo == "empresa":
            nombre = nombre_raw
        else:
            nombre = f"{apellido} {nombre_raw}".strip() if apellido and nombre_raw else (apellido or nombre_raw)
        if not nombre or nombre in ("None None", "None"):
            continue

        # Limpiar DNI: solo dígitos de 6+ chars
        import re
        dni = dni_raw if re.match(r"^\d{6,}$", dni_raw) else ""

        # Limpiar teléfono/WA: solo si tiene dígitos
        telefono = wa if re.search(r"\d{6,}", wa) else ""

        # Upsert
        existente = None
        if dni:
            existente = db.query(Cliente).filter(Cliente.dni_cuit == dni).first()
        if not existente:
            existente = db.query(Cliente).filter(
                text("LOWER(nombre) = LOWER(:n)")
            ).params(n=nombre).first()

        if existente:
            existente.tipo_cliente = tipo
            if dni:
                existente.dni_cuit = dni
            if telefono:
                existente.telefono = telefono
            if correo and correo != "None":
                existente.correo = correo
            actualizados += 1
        else:
            db.add(Cliente(
                nombre=nombre,
                dni_cuit=dni,
                telefono=telefono,
                correo=correo if correo != "None" else "",
                tipo_cliente=tipo,
                saldo_deudor=0,
                activo=True,
            ))
            creados += 1

        if (creados + actualizados) % 100 == 0:
            db.flush()

    db.commit()
    return {"ok": True, "creados": creados, "actualizados": actualizados}
